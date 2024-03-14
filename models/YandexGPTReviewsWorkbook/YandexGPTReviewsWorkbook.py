from openpyxl.worksheet.worksheet import Worksheet
import openpyxl

from Workbook import Workbook
from models.YandexMarketProduct.YandexMarketProduct import YandexMarketProduct

class YandexGPTReviewsWorkbook(Workbook):
    FILE_PATH = "C:/Users/user10/Desktop/Stuff/Отзывы/Yandex Market/Выводы от YandexGPT.xlsx"

    PRODUCT_ID_COL_TITLE = "product_id"
    PROS_COL_TITLE = "plus"
    CONS_COL_TITLE = "minus"

    def __init__(self):
        self._workbook = openpyxl.load_workbook(self.FILE_PATH)
        self._sheet: Worksheet = self._workbook.active

    def write_result(self, product: YandexMarketProduct):
        self._write_review(product=product)
        self._workbook.save(self.FILE_PATH)
    
    def _write_review(self, product: YandexMarketProduct):
        product_id_col_index = self.find_column_index(column_name=self.PRODUCT_ID_COL_TITLE)
        article_col_index = self.find_column_index(column_name=self.ARTICLE_COL_TITLE)
        pros_col_index = self.find_column_index(column_name=self.PROS_COL_TITLE)
        cons_col_index = self.find_column_index(column_name=self.CONS_COL_TITLE)

        self._sheet.cell(row=self._get_row_start_position(), column=product_id_col_index).value = product.id
        self._sheet.cell(row=self._get_row_start_position(), column=article_col_index).value = product.article
        self._sheet.cell(row=self._get_row_start_position(), column=pros_col_index).value = product.review._pros
        self._sheet.cell(row=self._get_row_start_position(), column=cons_col_index).value = product.review._cons
    
    def _get_row_start_position(self):
        return self._sheet.max_row + 1