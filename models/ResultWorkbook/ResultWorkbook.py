from typing import List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from Workbook import Workbook

from models.YandexMarketProduct.YandexMarketProduct import YandexMarketProduct
from models.Review.Review import Review

class ResultWorkbook(Workbook):
    RESULT_WORKBOOK_FILE_PATH = "C:/Users/user10/Desktop/Stuff/Отзывы/Yandex Market/Новые отзывы (Яндекс Маркет).xlsx"

    PRODUCT_ID_COL_TITLE = "product_id"
    ARTICLE_COL_TITLE = "c:vendor_sku"
    CUSTOMER_NAME_COL_TITLE = "customer_name"
    COMMENT_COL_TITLE = "review"
    RATE_COL_TITLE = "rate"
    PROS_COL_TITLE = "pros"
    CONS_COL_TITLE = "cons"

    def __init__(self):
        self._workbook = openpyxl.load_workbook(self.RESULT_WORKBOOK_FILE_PATH)
        self._sheet: Worksheet = self._workbook.active
        # self._columns_to_write = [
        #     self.PRODUCT_ID_COL_TITLE,
        #     self.ARTICLE_COL_TITLE,
        #     self.CUSTOMER_NAME_COL_TITLE,
        #     self.COMMENT_COL_TITLE,
        #     self.RATE_COL_TITLE
        # ]
        # self._columns_indexes_to_write = self.find_column_indexes(column_names=self._columns_to_write)

    def write_result(self, product: YandexMarketProduct):
        row_start_position = self._get_row_start_position()

        for index, review in enumerate(product.reviews):
            row_index = index + row_start_position

            self._write_review(row_index=row_index, review=review, product=product)

        self._workbook.save(self.RESULT_WORKBOOK_FILE_PATH)

    def _write_review(self, row_index: int, review: Review, product: YandexMarketProduct):
        product_id_col_index = self.find_column_index(column_name=self.PRODUCT_ID_COL_TITLE)
        article_col_index = self.find_column_index(column_name=self.ARTICLE_COL_TITLE)
        customer_name_col_index = self.find_column_index(column_name=self.CUSTOMER_NAME_COL_TITLE)
        comment_col_index = self.find_column_index(column_name=self.COMMENT_COL_TITLE)
        rate_col_index = self.find_column_index(column_name=self.RATE_COL_TITLE)
        pros_col_index = self.find_column_index(column_name=self.RATE_COL_TITLE)
        cons_col_index = self.find_column_index(column_name=self.RATE_COL_TITLE)

        self._sheet.cell(row=row_index, column=product_id_col_index).value = product.id
        self._sheet.cell(row=row_index, column=article_col_index).value = product.article

        self._sheet.cell(row=row_index, column=customer_name_col_index).value = review._customer_name
        self._sheet.cell(row=row_index, column=comment_col_index).value = review._comment
        self._sheet.cell(row=row_index, column=rate_col_index).value = review._rate
        self._sheet.cell(row=row_index, column=pros_col_index).value = review._pros
        self._sheet.cell(row=row_index, column=cons_col_index).value = review._cons

    def _get_row_start_position(self):
        return self._sheet.max_row + 1